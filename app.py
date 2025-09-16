import tkinter as tk
from tkinter import ttk, filedialog
from openpyxl import load_workbook
import ttkbootstrap as ttk
from ttkbootstrap.constants import *
import time
import threading
import pdfplumber
import os
import re
import pandas as pd
from pathlib import Path
import shutil
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
import zipfile
from tkinter import messagebox

# === INICIO DO PROGRAMA ===
print("-"*30)
print("✅ Função executada com sucesso!")
print("-"*30)
# ======= Exportar como PASTA (sem ZIP) e apagar TUDO criado =======
def _copytree_unique(src_dir, dest_dir_base):
    """
    Copia a pasta src_dir para dentro de dest_dir_base usando um nome que não conflite.
    Ex.: se 'relatorio' já existir, cria 'relatorio_2025-08-22_153012'.
    Retorna o caminho final copiado.
    """
    base_name = os.path.basename(src_dir.rstrip(os.sep)) or "relatorio"
    destino = os.path.join(dest_dir_base, base_name)

    if os.path.exists(destino):
        # cria sufixo com timestamp para evitar sobrescrita
        sufixo = time.strftime("%Y-%m-%d_%H%M%S")
        destino = os.path.join(dest_dir_base, f"{base_name}_{sufixo}")

    shutil.copytree(src_dir, destino)
    return destino

def _apagar_pastas_criadas():
    """
    Apaga TODAS as pastas criadas pelo código, incluindo 'relatorio'.
    """
    pastas_criadas = {
        "pdf",
        "processamento",
        "txt_limpo",
        "resultado",
        "ref_txt_limpo",
        "ref_resultado",
        "referencia",
        "relatorio",
    }
    for pasta in pastas_criadas:
        try:
            if os.path.isdir(pasta):
                shutil.rmtree(pasta)
                print(f"🗑️ Pasta removida: {pasta}")
        except Exception as e:
            print(f"❌ Erro ao remover pasta '{pasta}': {e}")

def exportar_relatorio():
    """Copia a pasta 'relatorio' para um diretório escolhido e depois apaga tudo que foi gerado pelo código."""
    global status_label, root

    try:
        if not os.path.isdir("relatorio"):
            messagebox.showwarning("Exportar", "A pasta 'relatorio' não existe ainda.")
            return

        # 🔴 REMOVER OS XLSX ANTES DA EXPORTAÇÃO
        _remover_resultados_relatorio()

        # Escolher uma PASTA de destino (não arquivo ZIP)
        destino_dir = filedialog.askdirectory(
            title="Escolha a pasta de destino para exportar 'relatorio'"
        )
        if not destino_dir:
            return  # usuário cancelou

        status_label.config(text="📦 Exportando pasta 'relatorio'...")
        root.update_idletasks()

        destino_final = _copytree_unique("relatorio", destino_dir)
        print(f"📁 Pasta 'relatorio' copiada para: {destino_final}")

        status_label.config(text="🧹 Apagando todas as pastas criadas pelo código...")
        root.update_idletasks()

        _apagar_pastas_criadas()

        status_label.config(text="✅ Exportação concluída e limpeza total realizada.")
        messagebox.showinfo(
            "Exportar",
            f"Exportação concluída!\n\nA pasta foi copiada para:\n{destino_final}\n\n"
            "Arquivos 'resultado_completo.xlsx' e 'resultado_unificado.xlsx' foram removidos antes da exportação.\n"
            "Todas as pastas geradas pelo código foram apagadas."
        )

    except Exception as e:
        status_label.config(text=f"❌ Erro ao exportar: {e}")
        messagebox.showerror("Exportar", f"Ocorreu um erro ao exportar:\n{e}")

def limpar_selecionados():
    """Apaga as pastas 'pdf' e 'referencia' e limpa as seleções/labels do app."""
    global arquivo_selecionado, referencia_selecionada
    global arquivo_paths, referencia_paths, arquivo_tipos, referencia_tipos
    global arquivo_label, referencia_label, status_label

    # Remover pastas criadas
    for pasta in ("pdf", "referencia"):
        try:
            if os.path.exists(pasta):
                shutil.rmtree(pasta)
                print(f"🗑️ Pasta removida: {pasta}")
        except Exception as e:
            print(f"❌ Erro ao remover '{pasta}': {e}")

    # Resetar estados de seleção
    arquivo_selecionado = False
    referencia_selecionada = False
    arquivo_paths.clear()
    referencia_paths.clear()
    arquivo_tipos.clear()
    referencia_tipos.clear()

    # Atualizar UI
    arquivo_label.config(text="📁 Nenhum arquivo selecionado")
    referencia_label.config(text="📋 Nenhuma referência selecionada")
    verificar_botoes()
    status_label.config(text="🧽 Seleções limpas e pastas apagadas.")

def _remover_resultados_relatorio():
    """
    Remove os arquivos resultado_completo.xlsx e resultado_unificado.xlsx
    de dentro da pasta 'relatorio', caso existam.
    """
    arquivos_para_remover = ["resultado_completo.xlsx", "resultado_unificado.xlsx"]
    for nome in arquivos_para_remover:
        caminho = os.path.join("relatorio", nome)
        if os.path.isfile(caminho):
            try:
                os.remove(caminho)
                print(f"🗑️ Arquivo removido de 'relatorio': {nome}")
            except Exception as e:
                print(f"❌ Erro ao remover '{nome}': {e}")

def apagar_pasta_relatorio():
    """
    Apaga a pasta 'relatorio' do diretório atual, se existir.
    """
    caminho_relatorio = os.path.join(os.getcwd(), "relatorio")
    if os.path.exists(caminho_relatorio) and os.path.isdir(caminho_relatorio):
        try:
            shutil.rmtree(caminho_relatorio)
            print("🗑️ Pasta 'relatorio' apagada com sucesso.")
        except Exception as e:
            print(f"❌ Erro ao apagar a pasta 'relatorio': {e}")
    else:
        print("ℹ️ Pasta 'relatorio' não encontrada.")

apagar_pasta_relatorio()

# Trata os arquivos PDF: Conversão para TXT e limpeza de txt e Conversão para XLSX (PRIMEIRA FUNÇÃO)
def tratar_pdf():
    print("Tratando PDF")
    PASTA_PDF = "pdf"
    PASTA_TXT = "processamento"
    PASTA_TXT_LIMPO = "txt_limpo"
    PASTA_XLSX = "resultado"

    os.makedirs(PASTA_TXT, exist_ok=True)
    os.makedirs(PASTA_TXT_LIMPO, exist_ok=True)
    os.makedirs(PASTA_XLSX, exist_ok=True)

    # =======================
    # EXTRAÇÃO DE TEXTO DO PDF
    # =======================
    def extrair_texto_pdf(arquivo_pdf, pasta_saida):
        print("Extrair Texto_pdf")
        # Extrai texto do PDF
        nome_base = os.path.splitext(os.path.basename(arquivo_pdf))[0]
        print(f"📄 Extraindo texto de '{arquivo_pdf}'...")
        texto_total = ""
        with pdfplumber.open(arquivo_pdf) as pdf:
            for pagina in pdf.pages:
                texto = pagina.extract_text()
                if texto:
                    texto_total += texto.strip() + "\n"

        # ⇩ NOVO: detectar empresa e renomear o TXT (se não for 'referência')
        EMPRESAS = [
            "EQUAGRIL EQUIPAMENTOS AGRICOLAS LTDA",
            "CAMBUCI METALURGICA LTDA",
            "MOTORTEM PECAS PARA MOTORES LTDA",
            "TVH BRASIL PECAS LTDA",
        ]

        is_referencia = ("referencia" in arquivo_pdf.lower()) or ("referência" in arquivo_pdf.lower())

        empresa_encontrada = None
        if not is_referencia:
            upper_all = texto_total.upper()
            for emp in EMPRESAS:
                if emp.upper() in upper_all:
                    empresa_encontrada = emp
                    break

        if empresa_encontrada:
            # Sanitiza para nome de arquivo
            safe_name = empresa_encontrada.replace("LTDA", "").strip()
            safe_name = safe_name.replace(" ", "_")
            safe_name = re.sub(r'[^A-Za-z0-9_]+', '_', safe_name)
            safe_name = re.sub(r'_+', '_', safe_name).strip('_')
            caminho_txt = os.path.join(pasta_saida, f"{safe_name}.txt")
            # Evitar sobrescrita
            i = 1
            while os.path.exists(caminho_txt):
                caminho_txt = os.path.join(pasta_saida, f"{safe_name}_{i}.txt")
                i += 1
            print(f"📝 Empresa detectada: {empresa_encontrada} → TXT salvo como: {os.path.basename(caminho_txt)}")
        else:
            caminho_txt = os.path.join(pasta_saida, f"{nome_base}.txt")

        # Salva
        with open(caminho_txt, "w", encoding="utf-8") as f:
            f.write(texto_total)

        print(f"✅ Texto extraído e salvo em: {caminho_txt}")
        return caminho_txt

    def processar_pdfs():
        print("Processar_pdfs")
        arquivos_pdf = [f for f in os.listdir(PASTA_PDF) if f.lower().endswith(".pdf")]
        # Criar pasta PDF de destino
        PASTA_PDF_DESTINO = "pdf"
        os.makedirs(PASTA_PDF_DESTINO, exist_ok=True)
        for nome_arquivo in arquivos_pdf:
            caminho_pdf = os.path.join(PASTA_PDF, nome_arquivo)
            try:
                extrair_texto_pdf(caminho_pdf, PASTA_TXT)
                # Copiar ou mover o arquivo PDF para a pasta "pdf"
                destino_pdf = os.path.join(PASTA_PDF_DESTINO, nome_arquivo)
                shutil.copy2(caminho_pdf, destino_pdf)  # ou use shutil.move() se quiser mover
            except Exception as e:
                print(f"❌ Erro ao processar '{nome_arquivo}': {e}")
        time.sleep(2)

    # LIMPEZA GLOBAL
    def limpeza_global(linha):
        print("Limpeza_global")
        if not linha.strip():
            return ""
        regex_palavras_espacadas = re.compile(r"\b(?:[A-Z]\s+){2,}[A-Z]\b")
        linha = regex_palavras_espacadas.sub(lambda m: m.group(0).replace(" ", ""), linha)
        linha = re.sub(r'(tractorcraft)([\w]+)', r'\1 \2', linha, flags=re.IGNORECASE)
        linha = re.sub(r'Item\s*Est[\.]?\s*Marca', 'Item Est Marca', linha, flags=re.IGNORECASE)
        linha = re.sub(r'(\d{8})\s+(\d+)\s+(\d+)\s+(\d+[,.]\d+)', r'\1 \2\3 \4', linha)
        linha = re.sub(r'\bANELDE\b', 'ANEL DE', linha)
        linha = re.sub(r'\bANELVE\s+DADOR\b', 'ANEL VEDADOR', linha)
        linha = re.sub(r'\bANELBO\s+RRACHAA\b', 'ANEL BORRACHA', linha)
        linha = re.sub(r'\bANELDI\s+STANCIADOR\b', 'ANEL DISTANCIADOR', linha)
        linha = re.sub(r'\bANELRE\s+TENTOR\b', 'ANEL RETENTOR', linha)
        linha = re.sub(r'\bANELDA\b', 'ANEL DA', linha)
        linha = re.sub(r'\bANELTR\s+AVA\b', 'ANEL TRAVA', linha)
        linha = re.sub(r'\bGARFOM\s+UDANCA\b', 'GARFO MUDANCA', linha)
        linha = re.sub(r'\bBOMBAH\s+IDR\b', 'BOMBA HIDR', linha)
        linha = re.sub(r'\bBOMBADA\b', 'BOMBA DA', linha)
        linha = re.sub(r'\bARR\s+RUELA\b', 'ARRUELA', linha)
        linha = re.sub(r'\bCORPODA\b', 'CORPO DA', linha)
        linha = re.sub(r'\bBRACODO\b', 'BRACO DO', linha)
        linha = re.sub(r'\bBUCHAP\s+INO\b', 'BUCHA PINO', linha)
        linha = re.sub(r'\bBUCHADO\b', 'BUCHA DO', linha)
        linha = re.sub(r'\bBUCHADA\b', 'BUCHA DA', linha)
        linha = re.sub(r'\bBUCHAE\s+IXO\b', 'BUCHA EIXO', linha)
        linha = re.sub(r'\bBUCHAS\s+UPERIOR\b', 'BUCHA SUPERIOR', linha)
        linha = re.sub(r'\bROLAMEN\s+TO\b', 'ROLAMENTO', linha)
        linha = re.sub(r'\bBUJAODA\b', 'BUJAO DA', linha)
        linha = re.sub(r'\bCALCODO PINHAO\b', 'CALCO DO PINHAO', linha)
        linha = re.sub(r'\bCALCOD\s+IFERENCAL\b', 'CALCO DIFRENCIAL', linha)
        linha = re.sub(r'\bCAPAPR\s+OTETORA\b', 'CAPA PROTETORA', linha)
        linha = re.sub(r'\bCAPAE\s+CUBO\b', 'CAPA DE CUBO', linha)
        linha = re.sub(r'\bCOLARDA\s+EMBREAGEM\b', 'COLAR DA EMBREAGEM', linha)
        linha = re.sub(r'\bCAIXADO\s+DIFERENCIA\b', 'CAIXA DO DIFERENCIAL', linha)
        linha = re.sub(r'\bCHAPADE\s+ENCOSTO\b', 'CHAPA DE ENCOSTO', linha)
        linha = re.sub(r'\bCREMALH\s+EIRA\b', 'CREMALHEIRA', linha)
        linha = re.sub(r'\bCOROAR\s+EDUTORA\b', 'COROA REDUTORA', linha)
        linha = re.sub(r'\bGUIADO\s+COLAR\b', 'GUIA DO COLAR', linha)
        linha = re.sub(r'\bCUBODA\b', 'CUBO DA', linha)
        linha = re.sub(r'\bTUBODA\b', 'TUBO DA', linha)
        linha = re.sub(r'\bTUBODO\b', 'TUBO DO', linha)
        linha = re.sub(r'\bTUBODE\b', 'TUBO DE', linha)
        linha = re.sub(r'\bDICOE\s+MBR\b', 'DISCO EMBR', linha)
        linha = re.sub(r'\bDISCOF\s+REIO\b', 'DISCO FREIO', linha)
        linha = re.sub(r'\bDISCODE\s+FREIO\b', 'DISCO DE FREIO', linha)
        linha = re.sub(r'\bDISCODE\s+BRONZE\b', 'DISCO DE BRONZE', linha)
        linha = re.sub(r'\bDISCO\s+DA\b', 'DISCO DA', linha)
        linha = re.sub(r'\bENGRENA\s+GEM\b', 'ENGRENAGEM', linha)
        linha = re.sub(r'\bELEMENTOP/\b', 'ELEMENTO P/', linha)
        linha = re.sub(r'\bINTERRU\s+PTOR\b', 'INTERRUPTOR', linha)
        linha = re.sub(r'\bCHAVEDE\b', 'CHAVE DE', linha)
        linha = re.sub(r'\bJUNTADO\b', 'JUNTA DO', linha)
        linha = re.sub(r'\bJUNTAC\s+ABECOTE\b', 'JUNTA CABECOTE', linha)
        linha = re.sub(r'\bJUNTADA\b', 'JUNTA DA', linha)
        linha = re.sub(r'\bLUVADA\b', 'LUVA DA', linha)
        linha = re.sub(r'\bLUVADE\b', 'LUVA DE', linha)
        linha = re.sub(r'\bMANGUEI\s+RA\b', 'MANGUEIRA', linha)
        linha = re.sub(r'\bMOLADA\b', 'MOLA DA', linha)
        linha = re.sub(r'\bMOLAFI\s+TA\b', 'MOLA FITA', linha)
        linha = re.sub(r'\bMOLASI\s+NCRONIZADO\b', 'MOLA SINCRONIZADO', linha)
        linha = re.sub(r'\bPINODO\b', 'PINO DO', linha)
        linha = re.sub(r'\bPINOEL\s+ASTICO\b', 'PINO ELASTICO', linha)
        linha = re.sub(r'\bPINODE\b', 'PINO DE', linha)
        linha = re.sub(r'\bPINOTR\s+AVA\b', 'PINO TRAVA', linha)
        linha = re.sub(r'\bPORCAC\s+ASTELO\b', 'PORCA CASTELO', linha)
        linha = re.sub(r'\bPORCAT\s+RAVA\b', 'PORCA TRAVA', linha)
        linha = re.sub(r'\bFIL\s+TRO\b', 'FILTRO', linha)
        linha = re.sub(r'\bP\s+R\s+E\s+-\s+FIL\s+TRO\b', 'PRE - FILTRO', linha)
        linha = re.sub(r'\bP\s+R\s+E\s+-\s+FILTRO\b', 'PRE - FILTRO', linha)
        linha = re.sub(r'\bHASTEDO\b', 'HASTE DO', linha)
        linha = re.sub(r'\bKIT\s+REP\s+ARO\b', 'KIT REPARO', linha)
        linha = re.sub(r'\bE\s+I\s+XO\s+EXTERNO\b', 'EIXO EXTERNO', linha)
        linha = re.sub(r'\bSILENCI\s+OSO\b', 'SILENCIOSO', linha)
        linha = re.sub(r'\bSOLENOI\s+DE\b', 'SOLENOIDE', linha)
        linha = re.sub(r'\bTAMPADO\b', 'TAMPA DO', linha)
        linha = re.sub(r'\bTAMPADA\b', 'TAMPA DA', linha)
        linha = re.sub(r'\bARTICUL\s+ACAO\b', 'ARTICULACAO', linha)
        linha = re.sub(r'\bTRATOME\s+TRO\b', 'TRATOMETRO', linha)
        linha = re.sub(r'\bTRAVAP\s+ARAFUSO\b', 'TRAVA PARAFUSO', linha)
        linha = re.sub(r'\bCANODE\b', 'CANO DE', linha)
        linha = re.sub(r'\bPOSICIO\s+NADOR\b', 'POSICIONADOR', linha)
        linha = re.sub(r'\bR\s+ETENTOR\b', 'RETENTOR', linha)
        linha = re.sub(r'\bR\s+ETENTOR\s+DO\b', 'RETENTOR DO', linha)
        linha = re.sub(r'\bR\s+ETENTOR\s+DA\b', 'RETENTOR DA', linha)
        linha = re.sub(r'\bR\s+ETENTOR\s+MANCAL\b', 'RETENTOR MANCAL', linha)
        linha = re.sub(r'\bE3NN9A660BA/1P\s+R\s+E\b', 'E3NN9A660BA/1 PRE', linha)
        linha = re.sub(r'\bE3NN9A660BA/2P\s+R\s+E\b', 'E3NN9A660BA/2 PRE', linha)
        linha = re.sub(r'\bE3NN9A660BA/3P\s+R\s+E\b', 'E3NN9A660BA/3 PRE', linha)
        linha = re.sub(r'\b(\d+)/(\d+)PRE\b', r'\1/\2 PRE', linha)
        linha = re.sub(r'\b([A-Z0-9]+)/(\d+)PRE\b', r'\1/\2 PRE', linha)
        linha = re.sub(r'\b60203502551/1R\s+ETENTOR\b', '60203502551/1RETENTOR', linha)
        linha = re.sub(r'\b60203502551/2R\s+ETENTOR\b', '60203502551/2RETENTOR', linha)
        linha = re.sub(r'\b60203502551ORR\s+ETENTOR\b', '60203502551ORRETENTOR', linha)
        linha = re.sub(r'\b60203502241/1R\s+ETENTOR\b', '60203502241/1RETENTOR', linha)
        linha = re.sub(r'\b60203502241/2R\s+ETENTOR\b', '60203502241/2RETENTOR', linha)
        linha = re.sub(r'\b60203502462/1R\s+ETENTOR\b', '60203502462/1RETENTOR', linha)
        linha = linha.replace('"', '')
        linha = re.sub(r'\b(\d+),00\b', r'\1', linha)
        linha = re.sub(r"\\*", "", linha)
        linha = linha.replace('*', '')
        # Correção de NCM quebrado em duas partes (padrões: '7 8902321' ou '4569801 3')
        def _corrigir_ncm_quebrado(txt):
            # Junta um dígito isolado com sete dígitos adjacentes formando 8 dígitos
            txt = re.sub(r'\b(\d)\s+(\d{7})\b', r'\1\2', txt)
            txt = re.sub(r'\b(\d{7})\s+(\d)\b', r'\1\2', txt)
            return txt
        linha = _corrigir_ncm_quebrado(linha)
        print(f"[DBG][global->ncm] {linha}")
        return linha

    # LIMPEZA CONDICIONAL POR CABEÇALHO
    def limpeza_condicional(linha, cabecalho):
        print("Limpeza_condicional")
        if cabecalho == "Item Quant. Código Descrição NCM R$ Unit. R$ Uni.St. R$ Uni.Ipi R$ Total":
            if ("Código" in linha and "Descrição" in linha and "R$ Unit." in linha):
                palavras = re.findall(r'\bCódigo\b|\bDescrição\b|R\$ Unit\.', linha)
                palavras = [p.replace('R$ Unit.', 'Valor Unitário') for p in palavras]
                partes = [p for p in palavras if p]
                if len(partes) > 1:
                    linha = partes[0] + '|' + ' '.join(partes[1:])
                elif partes:
                    linha = partes[0]
                return linha
        elif cabecalho == "Sq Referência Descrição NCM Qtde Vlr.Unit ICMS ST Vlr.IPI Vlr.Total":
            if linha.strip().startswith('Referência'):
                return "Código|Descrição|Valor Unitário"
            codigo_match = re.search(r'^\d+\s+(\S+)', linha)
            descricao_match = re.search(r'^\d+\s+\S+\s+(.+?)\s+\d{8}', linha)
            if not descricao_match:
                descricao_match = re.search(r'^\d+\s+\S+\s+(.+)', linha)
            valor_unitario_match = re.search(r'\d{8}\s+\d+\s+(\d+(?:,\d+)?)', linha)

            codigo = codigo_match.group(1) if codigo_match else ""
            descricao = descricao_match.group(1).strip() if descricao_match else ""
            valor_unitario = valor_unitario_match.group(1) if valor_unitario_match else ""

            if codigo or descricao:
                return f"{codigo}|{descricao}|{valor_unitario}"
            else:
                return linha
        elif cabecalho == "Seq. Identificação Código QtdeUMMarca Discriminação NCM ICMS Vr. IPI Vr. ST Pr. Unit. ** Total **":
            if linha.strip().startswith('Seq.'):
                return "Código|Descrição|Valor Unitário"
            codigo_match = re.search(r'^\d+\s+(\S+)', linha)
            codigo = codigo_match.group(1).strip() if codigo_match else ""
            descricao_match = re.search(r'^\d+\s+\S+\s+\S+\s+\S+\s+(.+?)\s+\d{8}', linha)
            if descricao_match:
                descricao_completa = descricao_match.group(1).strip()
                descricao = re.sub(r'^\S+\s+', '', descricao_completa)
            else:
                descricao = ""
            valor_unitario_match = re.search(r'(\d{1,3}(?:\.\d{3})*,\d+)\s+[\d.,]+$', linha)
            valor_unitario = valor_unitario_match.group(1) if valor_unitario_match else ""
            if codigo or descricao:
                return f"{codigo}|{descricao}|{valor_unitario}"
            else:
                return linha
        elif cabecalho == "ItemEst.Marca Código Descrição NCM CEST Un QtdPeso Unit. Vlr. Unit. Vlr. Total Prev. Entr % IPI% ICMS ICM Solid. Total c\\ Imp. Observação":
            if linha.strip().startswith('Item'):
                return "Código|Descrição|Valor Unitário"
            codigo_match = re.search(r'^\d+\s+\w\s+\w+\s+(\S+)', linha)
            codigo = codigo_match.group(1).strip() if codigo_match else ""
            descricao_match = re.search(r'^\d+\s+\w\s+\w+\s+\S+\s+(.+?)\s+\d{8}', linha)
            descricao = descricao_match.group(1).strip() if descricao_match else ""
            valor_unitario_match = re.search(r'\s+(\d+(?:,\d+)?)\s+\d+(?:,\d+)?\s+\d{2}/\d{2}/\d{4}', linha)
            valor_unitario = valor_unitario_match.group(1) if valor_unitario_match else ""
            if codigo or descricao:
                return f"{codigo}|{descricao}|{valor_unitario}"
            else:
                return linha
        return linha

    def inserir_cabecalho_manual(linhas_originais):
        for linha in linhas_originais:
            limpa = limpeza_global(linha)
            if "Descrição" in limpa or "Seq" in limpa:
                return limpa
        return None

    def limpar_txts():
        arquivos = [f for f in os.listdir(PASTA_TXT) if f.endswith('.txt')]
        regex_irrelevante = re.compile(
            r"(CNPJ|Telefone|Fax|Endere[cç]o|CEP|Email|Transportadora|Emiss[aã]o|Pagina|Página|Validade|Subtotal|Total|Condi[cç][aã]o|Impress[aã]o|Pedido|Vendedor|Contato|Natureza|Moeda|Inscri[cç][aã]o|Origem|Orçamento|Cliente : C.)",
            re.IGNORECASE
        )
        regex_item_valido = re.compile(r'\b\d{1,4}\b.*\b[\w/-]{3,}\b.*\b\d{1,4}[.,]\d{2}\b')
        regex_linha_quebrada = re.compile(r"^\s*[\w/-]+\s*$")

        # Empresas para buscar e renomear arquivos
        empresas = [
            "EQUAGRIL EQUIPAMENTOS AGRICOLAS LTDA",
            "CAMBUCI METALURGICA LTDA",
            "MOTORTEM PECAS PARA MOTORES LTDA",
            "TVH BRASIL PECAS LTDA"
        ]

        for nome_arquivo in arquivos:
            # Pular arquivos de referência
            if "referencia" in nome_arquivo.lower():
                print(f"⏭️ Pulando arquivo de referência: {nome_arquivo}")
                continue

            caminho_origem = os.path.join(PASTA_TXT, nome_arquivo)
            caminho_saida = os.path.join(PASTA_TXT_LIMPO, nome_arquivo)
            with open(caminho_origem, "r", encoding="utf-8") as f:
                linhas = f.readlines()
            if not linhas:
                continue
            # Detectar o cabeçalho do arquivo (literal)
            cabecalho = None
            for linha in linhas:
                linha_strip = linha.strip()
                if linha_strip == "Item Quant. Código Descrição NCM R$ Unit. R$ Uni.St. R$ Uni.Ipi R$ Total" or \
                   linha_strip == "Sq Referência Descrição NCM Qtde Vlr.Unit ICMS ST Vlr.IPI Vlr.Total" or \
                   linha_strip == "Seq. Identificação Código QtdeUMMarca Discriminação NCM ICMS Vr. IPI Vr. ST Pr. Unit. ** Total **" or \
                   linha_strip == "ItemEst.Marca Código Descrição NCM CEST Un QtdPeso Unit. Vlr. Unit. Vlr. Total Prev. Entr % IPI% ICMS ICM Solid. Total c\\ Imp. Observação":
                    cabecalho = linha_strip
                    break
            resultado = []
            linha_anterior = ""
            cabecalho_adicionado = False
            for linha in linhas:
                linha_strip = linha.strip()
                # Não remover linhas que começam com 'CÓD. DESCRIÇÃO'
                if linha_strip.startswith('CÓD. DESCRIÇÃO'):
                    resultado.append(linha_strip)
                    if not cabecalho_adicionado:
                        cabecalho_adicionado = True
                    continue
                # Remover linhas específicas pelo início
                if linha_strip.startswith('AV. PRESIDENTE DUTRA, 1907 - BRASILIA SOLICITAÇÂO DE COMPRA') or linha_strip.startswith('Rua Francisco Foga Dist.Ind.'):
                    continue
                # Remover linhas que começam com "FISCAL: " ou "Fiscal: "
                if linha_strip.startswith('FISCAL: ') or linha_strip.startswith('Fiscal: '):
                    continue
                # Remover linhas vazias ou linhas quebradas
                if not linha_strip or regex_linha_quebrada.match(linha_strip):
                    continue
                # Remover linhas irrelevantes
                if regex_irrelevante.search(linha_strip):
                    continue
                # Agora sim, aplique a limpeza global e condicional
                limpa = limpeza_global(linha_strip)
                print(f"[DBG][global] {limpa}")
                limpa = limpeza_condicional(limpa, cabecalho)
                print(f"[DBG][condicional] {limpa}")
                # Detectar e adicionar o cabeçalho apenas uma vez
                if not cabecalho_adicionado and ("Descrição" in limpa or "Seq" in limpa):
                    if linha_anterior:
                        resultado.append(linha_anterior)
                    resultado.append(limpa)
                    cabecalho_adicionado = True
                    continue
                # Adiciona se for linha de item (mais permissivo)
                if regex_item_valido.search(limpa) or (len(limpa.split()) >= 8 and any(c.isdigit() for c in limpa)) or (cabecalho in ["Sq Referência Descrição NCM Qtde Vlr.Unit ICMS ST Vlr.IPI Vlr.Total", "Seq. Identificação Código QtdeUMMarca Discriminação NCM ICMS Vr. IPI Vr. ST Pr. Unit. ** Total **", "ItemEst.Marca Código Descrição NCM CEST Un QtdPeso Unit. Vlr. Unit. Vlr. Total Prev. Entr % IPI% ICMS ICM Solid. Total c\\ Imp. Observação"] and "|" in limpa):
                    resultado.append(limpa)
                # Atualiza a linha anterior válida
                linha_anterior = limpa if not regex_irrelevante.search(limpa) else ""
            # Inserir manualmente o cabeçalho, se não foi adicionado
            if not cabecalho_adicionado:
                cabecalho_manual = inserir_cabecalho_manual(linhas)
                if cabecalho_manual:
                    resultado.insert(0, cabecalho_manual)
            # Remover a última linha se ela for muito diferente da penúltima ou antepenúltima em quantidade de palavras
            if len(resultado) >= 3:
                ultima = resultado[-1]
                penultima = resultado[-2]
                antepenultima = resultado[-3]
                len_ultima = len(ultima.split())
                len_penultima = len(penultima.split())
                len_antepenultima = len(antepenultima.split())
                if not (min(len_penultima, len_antepenultima) * 0.5 <= len_ultima <= max(len_penultima, len_antepenultima) * 2):
                    resultado.pop()
            # Excluir linhas em branco do resultado final (garantia extra)
            resultado = [linha for linha in resultado if linha.strip()]

            # Detectar empresa no conteúdo para renomear o arquivo
            empresa_encontrada = None
            conteudo_completo = "\n".join(resultado)

            for empresa in empresas:
                if empresa.upper() in conteudo_completo.upper():
                    empresa_encontrada = empresa
                    break

            # Renomear arquivo baseado na empresa encontrada
            if empresa_encontrada:
                # Criar novo nome baseado na empresa
                novo_nome = empresa_encontrada.replace(" ", "_").replace("LTDA", "").strip()
                extensao = os.path.splitext(nome_arquivo)[1]
                novo_nome_arquivo = f"{novo_nome}{extensao}"

                # Verificar se já existe arquivo com esse nome
                novo_caminho = os.path.join(PASTA_TXT_LIMPO, novo_nome_arquivo)
                contador = 1
                while os.path.exists(novo_caminho):
                    nome_base = os.path.splitext(novo_nome_arquivo)[0]
                    extensao = os.path.splitext(novo_nome_arquivo)[1]
                    novo_nome_arquivo = f"{nome_base}_{contador}{extensao}"
                    novo_caminho = os.path.join(PASTA_TXT_LIMPO, novo_nome_arquivo)
                    contador += 1

                # Atualizar caminho de saída com o novo nome
                caminho_saida = novo_caminho
                print(f"✅ Empresa encontrada: {empresa_encontrada}")
                print(f"🔄 Arquivo será salvo como: {novo_nome_arquivo}")

            # Salvar arquivo com o nome correto (original ou renomeado)
            with open(caminho_saida, "w", encoding="utf-8") as f_out:
                f_out.write("\n".join(resultado))

            # --- Limpeza condicional por cabeçalho após a limpeza global ---
            def limpeza_condicional_lote(linhas):
                import re
                if not linhas:
                    return linhas
                cabecalho = linhas[0].strip()
                if cabecalho == "Item Quant. Código Descrição NCM R$ Unit. R$ Uni.St. R$ Uni.Ipi R$ Total":
                    print("Acessando BLOCO 1")
                    linhas = [l for l in linhas if l.strip()]
                    novas_linhas = []
                    def limpar_apos_terceira_barra(linha):
                        partes = linha.split('|')
                        if len(partes) < 4:
                            return linha
                        nova_linha = '|'.join(partes[:3]) + '|'
                        resto = '|'.join(partes[3:])
                        m = re.search(r'\d+[,.]\d+', resto)
                        if m:
                            nova_linha += m.group(0)
                        return nova_linha
                    for i, linha in enumerate(linhas):
                        linha = re.sub(r'CARG8A4\s*8\s*21010', 'CARGA 84821010', linha)
                        if linha.strip().startswith('#'):
                            continue
                        linha = re.sub(r'/\d+', '', linha)
                        linha = re.sub(r'\b[\w]+/\b', '', linha)
                        linha = re.sub(r'^((?:[^|]*\|){1}[^|]*)\|', r'\1', linha)
                        if i == 0:
                            palavras = re.findall(r'\bCódigo\b|\bDescrição\b|R\$ Unit\.', linha)
                            palavras = [p.replace('R$ Unit.', 'Valor Unitário') for p in palavras]
                            partes = [p for p in palavras if p]
                            if len(partes) > 1:
                                linha_limpa = partes[0] + '|' + ' '.join(partes[1:])
                            elif partes:
                                linha_limpa = partes[0]
                            else:
                                linha_limpa = linha
                            linha_limpa = re.sub(r'(Descrição)(?!.*Descrição)', r'\1|', linha_limpa, count=1)
                            novas_linhas.append(linha_limpa)
                        else:
                            linha = re.sub(r'\b7\s+3\s+(\d{6,})\b', r'73\1', linha)
                            linha = re.sub(r'\b(\d)\s+(\d{6,})\b', r'\1\2', linha)
                            linha = re.sub(r'\b(\d)\s+(\d{7})\b', r'\1\2', linha)
                            m = re.search(r'(\w{6,}|\d{8})', linha)
                            if m:
                                start = m.start(1)
                                end = m.end(1)
                                linha_limpa = linha[:start] + '|' + linha[start:end] + '|' + linha[end:]
                                after = linha_limpa[end+2:]
                                m2 = re.search(r'(\d{8})', after)
                                if m2:
                                    pos = end + 2 + m2.start(1)
                                    linha_limpa = linha_limpa[:pos] + '|' + linha_limpa[pos:]
                                idx_barra = linha_limpa.find('|')
                                if idx_barra != -1:
                                    linha_limpa = linha_limpa[idx_barra:]
                                linha_limpa = limpar_apos_terceira_barra(linha_limpa)
                                m3 = re.match(r'\|([^|]*?/[^|]*?)(\s+|\|)(.*)', linha_limpa)
                                if m3:
                                    linha_limpa = '|' + m3.group(3)
                                if linha_limpa.startswith('|'):
                                    linha_limpa = linha_limpa[1:]
                                partes = linha_limpa.split('|')
                                if len(partes) > 2:
                                    primeiro = partes[0].strip()
                                    if primeiro == "504127326CNH" and "59,73" in partes[-1]:
                                        partes.insert(1, "GUARNICAOPARAEIXOROTAN")
                                        linha_limpa = '|'.join(partes)
                                    elif primeiro == "28027012" and "15,43" in partes[-1]:
                                        partes.insert(1, "PINO1070/TEMPERADO")
                                        linha_limpa = '|'.join(partes)
                                    elif primeiro == "9637611CNH" and "16,95" in partes[-1]:
                                        partes.insert(1, "PARAFUSODE ACOSEXTAVADO")
                                        linha_limpa = '|'.join(partes)
                                novas_linhas.append(linha_limpa)
                            else:
                                novas_linhas.append(linha)
                    return novas_linhas
                elif cabecalho == "Sq Referência Descrição NCM Qtde Vlr.Unit ICMS ST Vlr.IPI Vlr.Total":
                    print("Acessando BLOCO 2")
                    linhas = [l for l in linhas if l.strip()]
                    if linhas:
                        linhas[0] = "Código|Descrição|Valor Unitário"
                    linhas = [re.sub(r'/\d+', '', l) for l in linhas]
                    return linhas
                elif cabecalho == "Seq. Identificação Código QtdeUMMarca Discriminação NCM ICMS Vr. IPI Vr. ST Pr. Unit. ** Total **" or ("Seq." in cabecalho and "Identificação" in cabecalho and "QtdeUMMarca" in cabecalho and "Discriminação" in cabecalho):
                    print("Acessando BLOCO 3")
                    linhas = [l for l in linhas if l.strip()]
                    if linhas:
                        linhas[0] = "Código|Descrição|Valor Unitário"
                    if len(linhas) > 1:
                        linhas.pop(1)
                    linhas = [re.sub(r'/\d+', '', l) for l in linhas]
                    return linhas
                elif cabecalho == "ItemEst.Marca Código Descrição NCM CEST Un QtdPeso Unit. Vlr. Unit. Vlr. Total Prev. Entr % IPI% ICMS ICM Solid. Total c\\ Imp. Observação" or ("CEST" in cabecalho and "QtdPeso" in cabecalho) or cabecalho == "Código|Descrição|Valor Unitário":
                    print("Acessando BLOCO 4 - Item Est Marca")
                    linhas = [l for l in linhas if l.strip()]
                    if linhas:
                        linhas[0] = "Código|Descrição|Valor Unitário"
                    linhas = [re.sub(r'/\d+', '', l) for l in linhas]
                    return linhas
                return linhas

            with open(caminho_saida, "r", encoding="utf-8") as f:
                linhas_lidas = [l.rstrip('\n') for l in f]
            linhas_lidas = limpeza_condicional_lote(linhas_lidas)

            with open(caminho_saida, "w", encoding="utf-8") as f:
                f.write('\n'.join(linhas_lidas))
            print(f"✅ Arquivo limpo com cabeçalho salvo: {caminho_saida}")

    # CONVERSÃO TXT PARA XLSX
    def converter_txt_para_xlsx():
        """
        Converte todos os arquivos .txt da pasta txt_limpo para .xlsx na pasta xlsx
        """
        arquivos_txt = [f for f in os.listdir(PASTA_TXT_LIMPO) if f.endswith('.txt')]

        if not arquivos_txt:
            print("⚠️ Nenhum arquivo .txt encontrado na pasta txt_limpo.")
            return

        print(f"📊 Convertendo {len(arquivos_txt)} arquivo(s) TXT para XLSX...")

        for nome_arquivo in arquivos_txt:
            caminho_txt = os.path.join(PASTA_TXT_LIMPO, nome_arquivo)
            nome_xlsx = os.path.splitext(nome_arquivo)[0] + ".xlsx"
            caminho_xlsx = os.path.join(PASTA_XLSX, nome_xlsx)

            try:
                with open(caminho_txt, "r", encoding="utf-8") as f:
                    linhas = [linha.strip() for linha in f if linha.strip()]

                if not linhas:
                    print(f"⚠️ Arquivo vazio: {nome_arquivo}")
                    continue

                dados = []
                for linha in linhas:
                    if "|" in linha:
                        partes = linha.split("|")
                        while len(partes) < 3:
                            partes.append("")
                        dados.append(partes[:3])
                    else:
                        dados.append([linha, "", ""])

                if dados:
                    if dados[0] == ["Código", "Descrição", "Valor Unitário"] or "Código" in dados[0][0]:
                        colunas = ["Código", "Descrição", "Valor Unitário"]
                        df = pd.DataFrame(dados[1:], columns=colunas)
                    else:
                        colunas = ["Código", "Descrição", "Valor Unitário"]
                        df = pd.DataFrame(dados, columns=colunas)

                    df.to_excel(caminho_xlsx, index=False, engine='openpyxl')
                    print(f"✅ Convertido: {nome_arquivo} → {nome_xlsx}")
                else:
                    print(f"⚠️ Nenhum dado válido encontrado em: {nome_arquivo}")

            except Exception as e:
                print(f"❌ Erro ao converter '{nome_arquivo}': {e}")

        print(f"🎉 Conversão concluída! Arquivos salvos na pasta: {PASTA_XLSX}")

    # AJUSTE DE VALORES UNITÁRIOS
    def ajustar_valores_unitarios():
        # Justa valores unitários com 4 casas decimais para 2 casas decimais nos arquivos XLSX
        arquivos_xlsx = [f for f in os.listdir(PASTA_XLSX) if f.endswith('.xlsx')]
        if not arquivos_xlsx:
            print("⚠️ Nenhum arquivo .xlsx encontrado na pasta xlsx.")
            return
        print(f"🔧 Ajustando valores unitários em {len(arquivos_xlsx)} arquivo(s) XLSX...")
        for nome_arquivo in arquivos_xlsx:
            caminho_xlsx = os.path.join(PASTA_XLSX, nome_arquivo)
            try:
                df = pd.read_excel(caminho_xlsx, engine='openpyxl')
                if 'Valor Unitário' not in df.columns:
                    print(f"⚠️ Coluna 'Valor Unitário' não encontrada em: {nome_arquivo}")
                    continue
                valores_alterados = 0
                for i, valor in enumerate(df['Valor Unitário']):
                    if pd.isna(valor) or valor == "":
                        continue
                    valor_str = str(valor).strip()
                    match = re.match(r'^(\d+),(\d{4})$', valor_str)
                    if match:
                        parte_inteira = match.group(1)
                        parte_decimal = match.group(2)
                        nova_parte_decimal = parte_decimal[:2]
                        novo_valor = f"{parte_inteira},{nova_parte_decimal}"
                        df.at[i, 'Valor Unitário'] = novo_valor
                        valores_alterados += 1
                        print(f"  📝 {valor_str} → {novo_valor}")
                if valores_alterados > 0:
                    df.to_excel(caminho_xlsx, index=False, engine='openpyxl')
                    print(f"✅ Ajustado: {nome_arquivo} - {valores_alterados} valor(es) alterado(s)")
                else:
                    print(f"ℹ️ Nenhum ajuste necessário em: {nome_arquivo}")
            except Exception as e:
                print(f"❌ Erro ao ajustar '{nome_arquivo}': {e}")
        print(f"🎉 Ajuste de valores concluído!")
    def gerar_relatorio_menor():
        """
        Lê o arquivo resultado_unificado.xlsx, identifica itens com códigos repetidos,
        compara os valores unitários e mantém apenas o item com menor preço para cada código.
        """
        pasta_resultado = "resultado"
        arquivo_entrada = "resultado_unificado.xlsx"
        arquivo_saida = "resultado_menor_preco.xlsx"
        caminho_entrada = os.path.join(pasta_resultado, arquivo_entrada)
        caminho_saida = os.path.join(pasta_resultado, arquivo_saida)

        if not os.path.exists(caminho_entrada):
            print(f"❌ Arquivo {arquivo_entrada} não encontrado na pasta {pasta_resultado}.")
            return

        print(f"📊 Analisando arquivo {arquivo_entrada}...")

        try:
            df = pd.read_excel(caminho_entrada, engine='openpyxl')

            if 'Código' not in df.columns or 'Valor Unitário' not in df.columns:
                print("❌ Formato inválido: Colunas 'Código' ou 'Valor Unitário' não encontradas.")
                return

            def converter_valor(valor):
                if isinstance(valor, (int, float)):
                    return float(valor)
                if pd.isna(valor) or valor == "":
                    return float('inf')
                valor_str = str(valor).strip()
                valor_str = valor_str.replace('.', '').replace(',', '.')
                try:
                    return float(valor_str)
                except:
                    return float('inf')

            df['Valor_Numerico'] = df['Valor Unitário'].apply(converter_valor)

            print(f"🔍 Analisando {len(df)} itens em busca de códigos duplicados...")

            df_validos = df[df['Código'].notna() & (df['Código'] != "")]
            codigos_duplicados = df_validos[df_validos.duplicated(subset=['Código'], keep=False)]['Código'].unique()

            total_codigos = len(df_validos['Código'].unique())
            total_duplicados = len(codigos_duplicados)

            print(f"📋 Total de códigos únicos: {total_codigos}")
            print(f"🔄 Códigos duplicados encontrados: {total_duplicados}")

            df_final = df_validos.loc[df_validos.groupby('Código')['Valor_Numerico'].idxmin()]

            df_final = df_final.drop(columns=['Valor_Numerico'])
            df_final = df_final.sort_values('Código')

            df_final.to_excel(caminho_saida, index=False, engine='openpyxl')

            print(f"✅ Relatório de menor preço gerado com {len(df_final)} itens.")
            print(f"💰 {total_duplicados} códigos com múltiplos fornecedores foram analisados por preço.")
            print(f"📄 Arquivo salvo em: {caminho_saida}")

        except Exception as e:
            print(f"❌ Erro ao processar o arquivo: {e}")

    # EXECUÇÃO DAS ETAPAS
    processar_pdfs()
    limpar_txts()
    converter_txt_para_xlsx()
    ajustar_valores_unitarios()
    gerar_relatorio_menor()

# Trata arquivos: Referencia da Empresa (Conversão para TXT e limpeza de txt e Conversão para XLSX)
def tratar_referencias():
    """
    Função unificada para processar todos os arquivos PDF de referência.
    Converte PDF -> TXT limpo -> XLSX formatado
    """

    # Configuração de pastas
    PASTA_ENTRADA = "referencia"
    PASTA_SAIDA = "ref_txt_limpo"
    PASTA_XLSX = "ref_resultado"

    # 🧹 Limpar as pastas ref_txt_limpo e ref_resultado antes de começar
    for pasta in [PASTA_SAIDA, PASTA_XLSX]:
        if os.path.exists(pasta):
            shutil.rmtree(pasta)
        os.makedirs(pasta, exist_ok=True)
        print(f"🧼 Pasta limpa: {pasta}")

    def limpar_texto(texto):
        """
        Limpa e formata o texto extraído do PDF com regras avançadas
        """
        linhas = texto.splitlines()
        resultado = []
        cod_encontrado = False
        ignorar = True
        cabecalho_original = "CÓD. DESCRIÇÃO REFERÊNCIA REFERÊNCIA 2 QTDV.ENDEVDAOLROR"
        cabecalho_modificado = "NUM | DESCRIÇÃO | REFERÊNCIA | REFERÊNCIA 2 | QTD"
        padrao_item = re.compile(r"^.+$", re.UNICODE)

        for linha in linhas:
            linha_strip = linha.strip()

            if linha_strip.endswith('MM'):
                continue

            if "LTDA" in linha_strip.upper():
                continue

            if linha_strip == cabecalho_original:
                if not cod_encontrado:
                    resultado.append(cabecalho_modificado)
                    cod_encontrado = True
                ignorar = False
                continue

            if ignorar or not padrao_item.match(linha_strip):
                continue

            if re.match(r"^\d{2,3}\b", linha_strip):
                linha_strip = re.sub(r"^(\d{2,3})(\s+)", r"\1|\2", linha_strip)
            elif re.match(r"^\d\b", linha_strip):
                linha_strip = re.sub(r"^(\d)(\s+)", r"\1|\2", linha_strip)

            grupos = re.finditer(r"\b([A-Z0-9]{6,})\b", linha_strip, flags=re.IGNORECASE)
            for grupo in grupos:
                valor = grupo.group(1)
                if (
                    valor.isalpha()
                    or valor.upper().endswith("MM")
                    or re.match(r"^[A-Z]{2}\d{4}$", valor.upper())
                ):
                    continue
                inicio = grupo.start(1)
                linha_strip = linha_strip[:inicio] + "|" + linha_strip[inicio:]
                break

            linha_strip = re.sub(r"(\s)(\d+(?:,\d+)?)(\s*)$", r"\1|\2", linha_strip)

            if re.match(r"^\d{1,3}\|.+\|.*\|\d+(?:,\d+)?$", linha_strip):
                resultado.append(linha_strip)
        print("Referências Tratadas")
        return "\n".join(resultado)

    arquivos_processados = 0
    for nome_arquivo in os.listdir(PASTA_ENTRADA):
        if nome_arquivo.lower().endswith(".pdf"):
            caminho_pdf = os.path.join(PASTA_ENTRADA, nome_arquivo)
            nome_base = os.path.splitext(nome_arquivo)[0]
            caminho_txt = os.path.join(PASTA_SAIDA, f"{nome_base}.txt")
            caminho_xlsx = os.path.join(PASTA_XLSX, f"{nome_base}.xlsx")
            print(f"📄 Processando: {nome_arquivo}")
            try:
                with pdfplumber.open(caminho_pdf) as pdf:
                    texto_total = ""
                    for pagina in pdf.pages:
                        texto = pagina.extract_text()
                        if texto:
                            texto_total += texto + "\n"

                texto_limpo = limpar_texto(texto_total)

                with open(caminho_txt, "w", encoding="utf-8") as f:
                    f.write(texto_limpo)
                print(f"✅ TXT salvo: {caminho_txt}")

                with open(caminho_txt, 'r', encoding='utf-8') as f_txt:
                    linhas = [linha.strip().split('|') for linha in f_txt if linha.strip()]

                if linhas:
                    header = ["NUM", "DESCRIÇÃO", "REFERÊNCIA", "QTD"]
                    dados = []
                    for linha in linhas:
                        if len(linha) < 4:
                            linha.extend([''] * (4 - len(linha)))
                        elif len(linha) > 4:
                            linha = linha[:4]
                        dados.append([item.strip() for item in linha])

                    df = pd.DataFrame(dados, columns=header)

                    df.to_excel(caminho_xlsx, index=False)

                    wb = load_workbook(caminho_xlsx)
                    ws = wb.active
                    alinhamento_centro = Alignment(horizontal='center', vertical='center')
                    for row in ws.iter_rows():
                        for cell in row:
                            cell.alignment = alinhamento_centro
                    wb.save(caminho_xlsx)
                    print(f"✅ XLSX salvo: {caminho_xlsx}")
                    arquivos_processados += 1

            except Exception as e:
                print(f"❌ Erro ao processar '{nome_arquivo}': {e}")

    print(f"\n🎉 Processamento concluído!")
    print(f"📊 Total de arquivos processados: {arquivos_processados}")
    print(f"📁 TXT salvos em: {PASTA_SAIDA}")
    print(f"📁 XLSX salvos em: {PASTA_XLSX}")

# Tratar Sem Referência: Unificar arquivos Pdf em um Arquivo XLSX com Menores Preços Encontrados
def sem_referencia():
    print("INICIANDO PROCESSAMENTO SEM REFERÊNCIA")
    PASTA_PDF = "pdf"
    PASTA_TXT = "processamento"
    PASTA_TXT_LIMPO = "txt_limpo"
    PASTA_XLSX = "resultado"
    PASTA_RELATORIO = "relatorio"
    ARQUIVO_UNIFICADO = "resultado_unificado.xlsx"
    os.makedirs(PASTA_TXT, exist_ok=True)
    os.makedirs(PASTA_TXT_LIMPO, exist_ok=True)
    os.makedirs(PASTA_XLSX, exist_ok=True)
    os.makedirs(PASTA_RELATORIO, exist_ok=True)

    # COPIAR EXCELS SELECIONADOS PARA A PASTA "resultado"
    def copiar_excels_selecionados_para_resultado():
        """
        Copia quaisquer arquivos .xlsx que o usuário tenha selecionado para a pasta
        de saída PASTA_XLSX (resultado), para que sejam considerados na unificação.
        """
        try:
            # Usa a lista global de caminhos selecionados, quando existir
            global arquivo_paths
        except NameError:
            arquivo_paths = []

        copiados = 0
        if arquivo_paths:
            for caminho in arquivo_paths:
                try:
                    if isinstance(caminho, str) and caminho.lower().endswith('.xlsx'):
                        destino = os.path.join(PASTA_XLSX, os.path.basename(caminho))
                        # Evita sobrescrever se já existir um arquivo com o mesmo nome
                        if not os.path.exists(destino):
                            shutil.copy2(caminho, destino)
                            copiados += 1
                except Exception as e:
                    print(f"❌ Erro ao copiar Excel selecionado '{caminho}': {e}")
        print(f"📥 {copiados} Excel(s) copiado(s) para '{PASTA_XLSX}'.")

    # EXTRAÇÃO DE TEXTO DO PDF
    def extrair_texto_pdf(arquivo_pdf, pasta_saida):
        print("Extrair texto do PDF")
        nome_base = os.path.splitext(os.path.basename(arquivo_pdf))[0]
        texto_total = ""
        with pdfplumber.open(arquivo_pdf) as pdf:
            for pagina in pdf.pages:
                texto = pagina.extract_text()
                if texto:
                    texto_total += texto.strip() + "\n"
        # ⇩ NOVO: detectar empresa e renomear o TXT (se não for 'referência')
        EMPRESAS = [
            "EQUAGRIL EQUIPAMENTOS AGRICOLAS LTDA",
            "CAMBUCI METALURGICA LTDA",
            "MOTORTEM PECAS PARA MOTORES LTDA",
            "TVH BRASIL PECAS LTDA",
        ]
        is_referencia = ("referencia" in arquivo_pdf.lower()) or ("referência" in arquivo_pdf.lower())
        empresa_encontrada = None
        if not is_referencia:
            upper_all = texto_total.upper()
            for emp in EMPRESAS:
                if emp.upper() in upper_all:
                    empresa_encontrada = emp
                    break
        if empresa_encontrada:
            safe_name = empresa_encontrada.replace("LTDA", "").strip()
            safe_name = safe_name.replace(" ", "_")
            safe_name = re.sub(r'[^A-Za-z0-9_]+', '_', safe_name)
            safe_name = re.sub(r'_+', '_', safe_name).strip('_')
            caminho_txt = os.path.join(pasta_saida, f"{safe_name}.txt")
            i = 1
            while os.path.exists(caminho_txt):
                caminho_txt = os.path.join(pasta_saida, f"{safe_name}_{i}.txt")
                i += 1
            print(f"📝 Empresa detectada: {empresa_encontrada} → TXT salvo como: {os.path.basename(caminho_txt)}")
        else:
            caminho_txt = os.path.join(pasta_saida, f"{nome_base}.txt")

        with open(caminho_txt, "w", encoding="utf-8") as f:
            f.write(texto_total)
        return caminho_txt

    def processar_pdfs():
        if not os.path.exists(PASTA_PDF):
            print(f"❌ Pasta de entrada '{PASTA_PDF}' não encontrada.")
            return
        arquivos_pdf = [f for f in os.listdir(PASTA_PDF) if f.lower().endswith(".pdf")]
        if not arquivos_pdf:
            print("⚠️ Nenhum arquivo PDF encontrado na pasta.")
            return
        for nome_arquivo in arquivos_pdf:
            caminho_pdf = os.path.join(PASTA_PDF, nome_arquivo)
            try:
                extrair_texto_pdf(caminho_pdf, PASTA_TXT)
            except Exception as e:
                print(f"❌ Erro ao processar '{nome_arquivo}': {e}")
        time.sleep(2)

    # LIMPEZA GLOBAL (apenas substituições)
    def limpeza_global(linha):
        if not linha.strip():
            return ""
        print("\n🔎 Linha original:", linha)  # <-- debug inicial
        
        regex_palavras_espacadas = re.compile(r"\b(?:[A-Z]\s+){2,}[A-Z]\b")
        linha = regex_palavras_espacadas.sub(lambda m: m.group(0).replace(" ", ""), linha)
        linha = re.sub(r'(tractorcraft)([\w]+)', r'\1 \2', linha, flags=re.IGNORECASE)
        linha = re.sub(r'Item\s*Est[\.]?\s*Marca', 'Item Est Marca', linha, flags=re.IGNORECASE)
        linha = re.sub(r'(\d{8})\s+(\d+)\s+(\d+)\s+(\d+[,.]\d+)', r'\1 \2\3 \4', linha)
        linha = re.sub(r'\bANELDE\b', 'ANEL DE', linha)
        linha = re.sub(r'\bANELVE\s+DADOR\b', 'ANEL VEDADOR', linha)
        linha = re.sub(r'\bANELBO\s+RRACHAA\b', 'ANEL BORRACHA', linha)
        linha = re.sub(r'\bANELDI\s+STANCIADOR\b', 'ANEL DISTANCIADOR', linha)
        linha = re.sub(r'\bANELRE\s+TENTOR\b', 'ANEL RETENTOR', linha)
        linha = re.sub(r'\bANELDA\b', 'ANEL DA', linha)
        linha = re.sub(r'\bANELTR\s+AVA\b', 'ANEL TRAVA', linha)
        linha = re.sub(r'\bGARFOM\s+UDANCA\b', 'GARFO MUDANCA', linha)
        linha = re.sub(r'\bBOMBAH\s+IDR\b', 'BOMBA HIDR', linha)
        linha = re.sub(r'\bBOMBADA\b', 'BOMBA DA', linha)
        linha = re.sub(r'\bARR\s+RUELA\b', 'ARRUELA', linha)
        linha = re.sub(r'\bCORPODA\b', 'CORPO DA', linha)
        linha = re.sub(r'\bBRACODO\b', 'BRACO DO', linha)
        linha = re.sub(r'\bBUCHAP\s+INO\b', 'BUCHA PINO', linha)
        linha = re.sub(r'\bBUCHADO\b', 'BUCHA DO', linha)
        linha = re.sub(r'\bBUCHADA\b', 'BUCHA DA', linha)
        linha = re.sub(r'\bBUCHAE\s+IXO\b', 'BUCHA EIXO', linha)
        linha = re.sub(r'\bBUCHAS\s+UPERIOR\b', 'BUCHA SUPERIOR', linha)
        linha = re.sub(r'\bROLAMEN\s+TO\b', 'ROLAMENTO', linha)
        linha = re.sub(r'\bBUJAODA\b', 'BUJAO DA', linha)
        linha = re.sub(r'\bBOCALDO\b', 'BOCAL DO', linha)
        linha = re.sub(r'\bDISCODA\b', 'DISCO DA', linha)
        linha = re.sub(r'\bDEJ UNTAS\b', 'DE JUNTAS', linha)
        linha = re.sub(r'\bJUNTAT AMPA\b', 'JUNTA TAMPA', linha)
        linha = re.sub(r'\bDISCOE MBR\b', 'DISCO EMBR', linha)
        linha = re.sub(r'\bDISCOE MB.\b', 'DISCO EMB.', linha)
        linha = re.sub(r'\bGUARNICAOPARAEIXOROTANA\b', 'GUARNICAO PARA EIXO ROTAN', linha)
        linha = re.sub(r'\bCONJ JUNTA\b', 'CONJUNTA', linha)
        linha = re.sub(r'\bPARAFUSODE\b', 'PARAFUSO DE', linha)
        linha = re.sub(r'\bPORCADA\b', 'PORCA DA', linha)
        linha = re.sub(r'\bCALCODO PINHAO\b', 'CALCO DO PINHAO', linha)
        linha = re.sub(r'\bCALCOD\s+IFERENCAL\b', 'CALCO DIFRENCIAL', linha)
        linha = re.sub(r'\bCAPAPR\s+OTETORA\b', 'CAPA PROTETORA', linha)
        linha = re.sub(r'\bCAPAE\s+CUBO\b', 'CAPA DE CUBO', linha)
        linha = re.sub(r'\bCOLARDA\s+EMBREAGEM\b', 'COLAR DA EMBREAGEM', linha)
        linha = re.sub(r'\bCAIXADO\s+DIFERENCIA\b', 'CAIXA DO DIFERENCIAL', linha)
        linha = re.sub(r'\bCHAPADE\s+ENCOSTO\b', 'CHAPA DE ENCOSTO', linha)
        linha = re.sub(r'\bCREMALH\s+EIRA\b', 'CREMALHEIRA', linha)
        linha = re.sub(r'\bCOROAR\s+EDUTORA\b', 'COROA REDUTORA', linha)
        linha = re.sub(r'\bGUIADO\s+COLAR\b', 'GUIA DO COLAR', linha)
        linha = re.sub(r'\bCUBODA\b', 'CUBO DA', linha)
        linha = re.sub(r'\bTUBODA\b', 'TUBO DA', linha)
        linha = re.sub(r'\bTUBODO\b', 'TUBO DO', linha)
        linha = re.sub(r'\bTUBODE\b', 'TUBO DE', linha)
        linha = re.sub(r'\bDICOE\s+MBR\b', 'DISCO EMBR', linha)
        linha = re.sub(r'\bDISCOF\s+REIO\b', 'DISCO FREIO', linha)
        linha = re.sub(r'\bDISCODE\s+FREIO\b', 'DISCO DE FREIO', linha)
        linha = re.sub(r'\bDISCODE\s+BRONZE\b', 'DISCO DE BRONZE', linha)
        linha = re.sub(r'\bDISCO\s+DA\b', 'DISCO DA', linha)
        linha = re.sub(r'\bENGRENA\s+GEM\b', 'ENGRENAGEM', linha)
        linha = re.sub(r'\bELEMENTOP/\b', 'ELEMENTO P/', linha)
        linha = re.sub(r'\bINTERRU\s+PTOR\b', 'INTERRUPTOR', linha)
        linha = re.sub(r'\bCHAVEDE\b', 'CHAVE DE', linha)
        linha = re.sub(r'\bJUNTADO\b', 'JUNTA DO', linha)
        linha = re.sub(r'\bJUNTAC\s+ABECOTE\b', 'JUNTA CABECOTE', linha)
        linha = re.sub(r'\bJUNTADA\b', 'JUNTA DA', linha)
        linha = re.sub(r'\bLUVADA\b', 'LUVA DA', linha)
        linha = re.sub(r'\bLUVADE\b', 'LUVA DE', linha)
        linha = re.sub(r'\bMANGUEI\s+RA\b', 'MANGUEIRA', linha)
        linha = re.sub(r'\bMOLADA\b', 'MOLA DA', linha)
        linha = re.sub(r'\bMOLAFI\s+TA\b', 'MOLA FITA', linha)
        linha = re.sub(r'\bMOLASI\s+NCRONIZADO\b', 'MOLA SINCRONIZADO', linha)
        linha = re.sub(r'\bPINODO\b', 'PINO DO', linha)
        linha = re.sub(r'\bPINOEL\s+ASTICO\b', 'PINO ELASTICO', linha)
        linha = re.sub(r'\bPINODE\b', 'PINO DE', linha)
        linha = re.sub(r'\bPINOTR\s+AVA\b', 'PINO TRAVA', linha)
        linha = re.sub(r'\bPORCAC\s+ASTELO\b', 'PORCA CASTELO', linha)
        linha = re.sub(r'\bPORCAT\s+RAVA\b', 'PORCA TRAVA', linha)
        linha = re.sub(r'\bFIL\s+TRO\b', 'FILTRO', linha)
        linha = re.sub(r'\bP\s+R\s+E\s+-\s+FIL\s+TRO\b', 'PRE - FILTRO', linha)
        linha = re.sub(r'\bP\s+R\s+E\s+-\s+FILTRO\b', 'PRE - FILTRO', linha)
        linha = re.sub(r'\bHASTEDO\b', 'HASTE DO', linha)
        linha = re.sub(r'\bKIT\s+REP\s+ARO\b', 'KIT REPARO', linha)
        linha = re.sub(r'\bE\s+I\s+XO\s+EXTERNO\b', 'EIXO EXTERNO', linha)
        linha = re.sub(r'\bSILENCI\s+OSO\b', 'SILENCIOSO', linha)
        linha = re.sub(r'\bSOLENOI\s+DE\b', 'SOLENOIDE', linha)
        linha = re.sub(r'\bTAMPADO\b', 'TAMPA DO', linha)
        linha = re.sub(r'\bTAMPADA\b', 'TAMPA DA', linha)
        linha = re.sub(r'\bARTICUL\s+ACAO\b', 'ARTICULACAO', linha)
        linha = re.sub(r'\bTRATOME\s+TRO\b', 'TRATOMETRO', linha)
        linha = re.sub(r'\bTRAVAP\s+ARAFUSO\b', 'TRAVA PARAFUSO', linha)
        linha = re.sub(r'\bCANODE\b', 'CANO DE', linha)
        linha = re.sub(r'\bPOSICIO\s+NADOR\b', 'POSICIONADOR', linha)
        linha = re.sub(r'\bR\s+ETENTOR\b', 'RETENTOR', linha)
        linha = re.sub(r'\bR\s+ETENTOR\s+DO\b', 'RETENTOR DO', linha)
        linha = re.sub(r'\bR\s+ETENTOR\s+DA\b', 'RETENTOR DA', linha)
        linha = re.sub(r'\bR\s+ETENTOR\s+MANCAL\b', 'RETENTOR MANCAL', linha)
        linha = re.sub(r'\bE3NN9A660BA/1P\s+R\s+E\b', 'E3NN9A660BA/1 PRE', linha)
        linha = re.sub(r'\bE3NN9A660BA/2P\s+R\s+E\b', 'E3NN9A660BA/2 PRE', linha)
        linha = re.sub(r'\bE3NN9A660BA/3P\s+R\s+E\b', 'E3NN9A660BA/3 PRE', linha)
        linha = re.sub(r'\b(\d+)/(\d+)PRE\b', r'\1/\2 PRE', linha)
        linha = re.sub(r'\b([A-Z0-9]+)/(\d+)PRE\b', r'\1/\2 PRE', linha)
        linha = re.sub(r'\b60203502551/1R\s+ETENTOR\b', '60203502551/1RETENTOR', linha)
        linha = re.sub(r'\b60203502551/2R\s+ETENTOR\b', '60203502551/2RETENTOR', linha)
        linha = re.sub(r'\b60203502551ORR\s+ETENTOR\b', '60203502551ORRETENTOR', linha)
        linha = re.sub(r'\b60203502241/1R\s+ETENTOR\b', '60203502241/1RETENTOR', linha)
        linha = re.sub(r'\b60203502241/2R\s+ETENTOR\b', '60203502241/2RETENTOR', linha)
        linha = re.sub(r'\b60203502462/1R\s+ETENTOR\b', '60203502462/1RETENTOR', linha)
        linha = linha.replace('"', '') # Removendo as aspas
        linha = re.sub(r'\b(\d+),00\b', r'\1', linha) # Removendo as vírgulas
        linha = re.sub(r"\\*", "", linha) # Removendo as barras
        linha = linha.replace('*', '') # Removendo as asteriscos
        def _corrigir_ncm_quebrado(txt): # Corrigindo o NCM quebrado
            txt = re.sub(r'\b(\d)\s+(\d{7})\b', r'\1\2', txt)
            txt = re.sub(r'\b(\d{7})\s+(\d)\b', r'\1\2', txt)
            return txt
        linha = _corrigir_ncm_quebrado(linha)
        return linha

    # LIMPEZA CONDICIONAL POR CABEÇALHO
    def limpeza_condicional(linha, cabecalho):
        print("Limpeza Condicional")
        if cabecalho == "Item Quant. Código Descrição NCM R$ Unit. R$ Uni.St. R$ Uni.Ipi R$ Total":
            print("------Equagril-------")
            if "Código" in linha and "Descrição" in linha:
                return "Código / Descrição / Valor Unitário"

            m = re.match(r'^\s*\d+\s+\d+\s+([A-Z0-9\/]+)\s+(.+?)\s+(\d+,\d{2})', linha)
            if m:
                codigo = m.group(1)
                descricao = m.group(2).strip()
                valor = m.group(3)
                print(f"Match({linha}): {codigo} | {descricao} | {valor}")
                return f"{codigo} | {descricao} | {valor}"
            return ""  
        elif cabecalho == "Sq Referência Descrição NCM Qtde Vlr.Unit ICMS ST Vlr.IPI Vlr.Total":
            if linha.strip().startswith('Referência'):
                return "Código|Descrição|Valor Unitário"
            codigo_match = re.search(r'^\d+\s+(\S+)', linha)
            descricao_match = re.search(r'^\d+\s+\S+\s+(.+?)\s+\d{8}', linha)
            if not descricao_match:
                descricao_match = re.search(r'^\d+\s+\S+\s+(.+)', linha)
            valor_unitario_match = re.search(r'\d{8}\s+\d+\s+(\d+(?:,\d+)?)', linha)

            codigo = codigo_match.group(1) if codigo_match else ""
            descricao = descricao_match.group(1).strip() if descricao_match else ""
            valor_unitario = valor_unitario_match.group(1) if valor_unitario_match else ""

            if codigo or descricao:
                return f"{codigo}|{descricao}|{valor_unitario}"
            else:
                return linha
        elif cabecalho == "Seq. Identificação Código QtdeUMMarca Discriminação NCM ICMS Vr. IPI Vr. ST Pr. Unit. ** Total **":
            if linha.strip().startswith('Seq.'):
                return "Código|Descrição|Valor Unitário"
            codigo_match = re.search(r'^\d+\s+(\S+)', linha)
            codigo = codigo_match.group(1).strip() if codigo_match else ""
            descricao_match = re.search(r'^\d+\s+\S+\s+\S+\s+\S+\s+(.+?)\s+\d{8}', linha)
            if descricao_match:
                descricao_completa = descricao_match.group(1).strip()
                descricao = re.sub(r'^\S+\s+', '', descricao_completa)
            else:
                descricao = ""
            valor_unitario_match = re.search(r'(\d{1,3}(?:\.\d{3})*,\d+)\s+[\d.,]+$', linha)
            valor_unitario = valor_unitario_match.group(1) if valor_unitario_match else ""
            if codigo or descricao:
                return f"{codigo}|{descricao}|{valor_unitario}"
            else:
                return linha
        elif cabecalho == "ItemEst.Marca Código Descrição NCM CEST Un QtdPeso Unit. Vlr. Unit. Vlr. Total Prev. Entr % IPI% ICMS ICM Solid. Total c\\ Imp. Observação":
            if linha.strip().startswith('Item'):
                return "Código|Descrição|Valor Unitário"
            codigo_match = re.search(r'^\d+\s+\w\s+\w+\s+(\S+)', linha)
            codigo = codigo_match.group(1).strip() if codigo_match else ""
            descricao_match = re.search(r'^\d+\s+\w\s+\w+\s+\S+\s+(.+?)\s+\d{8}', linha)
            descricao = descricao_match.group(1).strip() if descricao_match else ""
            valor_unitario_match = re.search(r'\s+(\d+(?:,\d+)?)\s+\d+(?:,\d+)?\s+\d{2}/\d{2}/\d{4}', linha)
            valor_unitario = valor_unitario_match.group(1) if valor_unitario_match else ""
            if codigo or descricao:
                return f"{codigo}|{descricao}|{valor_unitario}"
            else:
                return linha
        return linha

    def limpar_txts():
        arquivos = [f for f in os.listdir(PASTA_TXT) if f.endswith('.txt')]

        empresas = [
            "EQUAGRIL EQUIPAMENTOS AGRICOLAS LTDA",
            "CAMBUCI METALURGICA LTDA",
            "MOTORTEM PECAS PARA MOTORES LTDA",
            "TVH BRASIL PECAS LTDA"
        ]

        for nome_arquivo in arquivos:
            if "referencia" in nome_arquivo.lower():
                print(f"⏭️ Pulando arquivo de referência: {nome_arquivo}")
                continue

            caminho_origem = os.path.join(PASTA_TXT, nome_arquivo)
            caminho_saida = os.path.join(PASTA_TXT_LIMPO, nome_arquivo)

            with open(caminho_origem, "r", encoding="utf-8") as f:
                linhas = f.readlines()

            if not linhas:
                continue

            # Detectar cabeçalho
            cabecalho = None
            for linha in linhas:
                linha_strip = linha.strip()
                if linha_strip in [
                    "Item Quant. Código Descrição NCM R$ Unit. R$ Uni.St. R$ Uni.Ipi R$ Total",
                    "Sq Referência Descrição NCM Qtde Vlr.Unit ICMS ST Vlr.IPI Vlr.Total",
                    "Seq. Identificação Código QtdeUMMarca Discriminação NCM ICMS Vr. IPI Vr. ST Pr. Unit. ** Total **",
                    "ItemEst.Marca Código Descrição NCM CEST Un QtdPeso Unit. Vlr. Unit. Vlr. Total Prev. Entr % IPI% ICMS ICM Solid. Total c\\ Imp. Observação"
                ]:
                    cabecalho = linha_strip
                    break

            resultado = []

            # Força o cabeçalho correto no início
            if cabecalho:
                resultado.append("Código | Descrição | Valor Unitário")

            # Processar cada linha
            for linha in linhas:
                linha_strip = linha.strip()
                if not linha_strip:
                    continue

                limpa = limpeza_global(linha_strip)
                limpa = limpeza_condicional(limpa, cabecalho)

                # Salvar somente se limpeza_condicional trouxe código|descrição|valor
                if limpa and "|" in limpa:
                    resultado.append(limpa)

            # Garantir que não salva vazio
            if len(resultado) > 1:
                with open(caminho_saida, "w", encoding="utf-8") as f_out:
                    f_out.write("\n".join(resultado))

                print(f"✅ Arquivo limpo salvo: {caminho_saida}")
            else:
                print(f"⚠️ Nenhum item válido encontrado em {nome_arquivo}")


    # CONVERSÃO TXT PARA XLSX
    def converter_txt_para_xlsx():
        """
        Converte todos os arquivos .txt da pasta txt_limpo para .xlsx na pasta xlsx
        """
        arquivos_txt = [f for f in os.listdir(PASTA_TXT_LIMPO) if f.endswith('.txt')]

        if not arquivos_txt:
            print("⚠️ Nenhum arquivo .txt encontrado na pasta txt_limpo.")
            return
        print(f"📊 Convertendo {len(arquivos_txt)} arquivo(s) TXT para XLSX...")
        for nome_arquivo in arquivos_txt:
            caminho_txt = os.path.join(PASTA_TXT_LIMPO, nome_arquivo)
            nome_xlsx = os.path.splitext(nome_arquivo)[0] + ".xlsx"
            caminho_xlsx = os.path.join(PASTA_XLSX, nome_xlsx)
            try:
                with open(caminho_txt, "r", encoding="utf-8") as f:
                    linhas = [linha.strip() for linha in f if linha.strip()]
                if not linhas:
                    print(f"⚠️ Arquivo vazio: {nome_arquivo}")
                    continue

                dados = []
                for linha in linhas:
                    if "|" in linha:
                        partes = linha.split("|")
                        while len(partes) < 3:
                            partes.append("")
                        dados.append(partes[:3])
                    else:
                        dados.append([linha, "", ""])

                if dados:
                    if dados[0] == ["Código", "Descrição", "Valor Unitário"] or "Código" in dados[0][0]:
                        colunas = ["Código", "Descrição", "Valor Unitário"]
                        df = pd.DataFrame(dados[1:], columns=colunas)
                    else:
                        colunas = ["Código", "Descrição", "Valor Unitário"]
                        df = pd.DataFrame(dados, columns=colunas)
                    df.to_excel(caminho_xlsx, index=False, engine='openpyxl')
                    print(f"✅ Convertido: {nome_arquivo} → {nome_xlsx}")
                else:
                    print(f"⚠️ Nenhum dado válido encontrado em: {nome_arquivo}")
            except Exception as e:
                print(f"❌ Erro ao converter '{nome_arquivo}': {e}")
        print(f"🎉 Conversão concluída! Arquivos salvos na pasta: {PASTA_XLSX}")

    # AJUSTE DE VALORES UNITÁRIOS
    def ajustar_valores_unitarios():
        """
        Ajusta valores unitários com 4 casas decimais para 2 casas decimais nos arquivos XLSX
        e remove os traços '-' da coluna Código.
        """
        arquivos_xlsx = [f for f in os.listdir(PASTA_XLSX) if f.endswith('.xlsx')]

        if not arquivos_xlsx:
            print("⚠️ Nenhum arquivo .xlsx encontrado na pasta xlsx.")
            return
        print(f"🔧 Ajustando valores e códigos em {len(arquivos_xlsx)} arquivo(s) XLSX...")

        for nome_arquivo in arquivos_xlsx:
            caminho_xlsx = os.path.join(PASTA_XLSX, nome_arquivo)
            try:
                df = pd.read_excel(caminho_xlsx, engine='openpyxl')

                alteracoes = False
                valores_alterados = 0
                codigos_alterados = 0

                if 'Valor Unitário' in df.columns:
                    for i, valor in enumerate(df['Valor Unitário']):
                        if pd.isna(valor) or valor == "":
                            continue
                        valor_str = str(valor).strip()
                        match = re.match(r'^(\d+),(\d{4})$', valor_str)
                        if match:
                            parte_inteira = match.group(1)
                            parte_decimal = match.group(2)[:2]
                            novo_valor = f"{parte_inteira},{parte_decimal}"
                            df.at[i, 'Valor Unitário'] = novo_valor
                            valores_alterados += 1
                            alteracoes = True
                            print(f"  📝 {valor_str} → {novo_valor}")
                else:
                    print(f"⚠️ Coluna 'Valor Unitário' não encontrada em: {nome_arquivo}")

                if 'Código' in df.columns:
                    codigos_antigos = df['Código'].astype(str)
                    df['Código'] = codigos_antigos.str.replace(r'[-/]', '', regex=True)
                    codigos_alterados = sum(c1 != c2 for c1, c2 in zip(codigos_antigos, df['Código']))
                    if codigos_alterados > 0:
                        alteracoes = True
                        print(f"  🔧 {codigos_alterados} código(s) ajustado(s) (removido '-' e '/')")

                if alteracoes:
                    df.to_excel(caminho_xlsx, index=False, engine='openpyxl')
                    print(f"✅ Arquivo atualizado: {nome_arquivo}")
                else:
                    print(f"ℹ️ Nenhuma alteração necessária em: {nome_arquivo}")

            except Exception as e:
                print(f"❌ Erro ao ajustar '{nome_arquivo}': {e}")

        print(f"🎉 Ajuste de valores e códigos concluído!")

    # UNIFICAÇÃO DE ARQUIVOS XLSX
    def unificar_arquivos_xlsx():

        arquivos_xlsx = [f for f in os.listdir(PASTA_XLSX) if f.endswith('.xlsx')]
        if not arquivos_xlsx:
            print("⚠️ Nenhum arquivo .xlsx encontrado na pasta resultado.")
            return
        print(f"🔄 Unificando {len(arquivos_xlsx)} arquivo(s) XLSX...")
        dfs_unidos = []
        for nome_arquivo in arquivos_xlsx:
            caminho_xlsx = os.path.join(PASTA_XLSX, nome_arquivo)
            try:
                df = pd.read_excel(caminho_xlsx, engine='openpyxl')
                if 'Valor Unitário' not in df.columns:
                    print(f"⚠️ Coluna 'Valor Unitário' não encontrada em: {nome_arquivo}")
                    continue
                fornecedor = Path(nome_arquivo).stem
                df['Fornecedor'] = fornecedor
                colunas = ['Fornecedor'] + [col for col in df.columns if col != 'Fornecedor']
                df = df[colunas]
                dfs_unidos.append(df)
                print(f"✅ Adicionado: {nome_arquivo} - {len(df)} linhas")
            except Exception as e:
                print(f"❌ Erro ao processar '{nome_arquivo}': {e}")
        if not dfs_unidos:
            print("⚠️ Nenhum dado válido para unificação.")
            return
        df_unificado = pd.concat(dfs_unidos, ignore_index=True)
        # Salva o relatório unificado (todos os itens concatenados)
        caminho_relatorio_unificado = os.path.join(PASTA_RELATORIO, "relatorio_unificado.xlsx")
        df_unificado.to_excel(caminho_relatorio_unificado, index=False, engine='openpyxl')
        print(f"📊 Relatório unificado salvo em: {caminho_relatorio_unificado} ({len(df_unificado)} itens)")
        # Mantém compatibilidade com nome antigo
        caminho_completo = os.path.join(PASTA_RELATORIO, "resultado_completo.xlsx")
        df_unificado.to_excel(caminho_completo, index=False, engine='openpyxl')
        try:
            df_unificado['temp_valor'] = pd.to_numeric(
                df_unificado['Valor Unitário']
                .astype(str)
                .str.replace('.', '', regex=False)
                .str.replace(',', '.', regex=False),
                errors='coerce'
            )
            nao_numericos = df_unificado[df_unificado['temp_valor'].isna()]
            if len(nao_numericos) > 0:
                print(f"⚠️ {len(nao_numericos)} itens com valores não numéricos encontrados")
                for _, row in nao_numericos.iterrows():
                    print(f"  - Código: {row['Código']}, Valor: '{row['Valor Unitário']}'")
                df_unificado = df_unificado[~df_unificado['temp_valor'].isna()].copy()
            if not df_unificado.empty:
                idx_min = df_unificado.groupby(['Código'])['temp_valor'].idxmin()
                df_unificado = df_unificado.loc[idx_min].reset_index(drop=True)
                df_unificado.drop('temp_valor', axis=1, inplace=True)
                print(f"🧹 Itens duplicados removidos. Mantidos apenas os de menor valor para cada código.")
            else:
                print("⚠️ Nenhum item com valor numérico válido encontrado para processamento.")
        except Exception as e:
            print(f"❌ Erro ao filtrar duplicados: {e}")
        # Salva o resultado comparado (deduplicado pelo menor valor por Código)
        caminho_resultado_comparado = os.path.join(PASTA_RELATORIO, "resultado_comparado.xlsx")
        df_unificado.to_excel(caminho_resultado_comparado, index=False, engine='openpyxl')
        print(f"🎉 Comparação concluída! {len(df_unificado)} itens únicos (menor valor por código).")
        print(f"📊 Arquivo de resultado comparado salvo em: {caminho_resultado_comparado}")
        # Mantém compatibilidade com nome antigo esperado por outras rotinas
        caminho_unificado = os.path.join(PASTA_RELATORIO, ARQUIVO_UNIFICADO)
        df_unificado.to_excel(caminho_unificado, index=False, engine='openpyxl')

    # EXECUÇÃO DAS ETAPAS
    processar_pdfs()
    limpar_txts()
    converter_txt_para_xlsx()
    # Garante incluir Excel(s) já selecionados pelo usuário na unificação
    copiar_excels_selecionados_para_resultado()
    ajustar_valores_unitarios()
    unificar_arquivos_xlsx()

# Pega dados da Coluna Código (Arquivo Referência) e Puxa do Arquivo XLSX (Arquivos Sem Referência Unificado)
def busca_codigo():
    def converter_valor(valor_str):
        try:
            valor_str = str(valor_str).strip()
            valor_str = valor_str.replace('.', '').replace(',', '.')
            valor_float = float(valor_str)
            valor_corrigido = float(re.match(r'^(\d+\.\d{1,2})', f"{valor_float:.4f}").group(1))
            return valor_corrigido
        except:
            return None

    def encontrar_arquivo_referencia(pasta):
        """Retorna o primeiro arquivo .xlsx encontrado na pasta"""
        for nome in os.listdir(pasta):
            if nome.lower().endswith(".xlsx"):
                return os.path.join(pasta, nome)
        return None

    def comparar_codigos_e_gerar_resultado():
        PASTA_REFERENCIA = 'ref_resultado'
        ARQ_REFERENCIA = encontrar_arquivo_referencia(PASTA_REFERENCIA)
        ARQ_UNIFICADO = 'relatorio/resultado_unificado.xlsx'
        ARQ_SAIDA = 'relatorio/resultado_comparado.xlsx'
        if not ARQ_REFERENCIA or not os.path.exists(ARQ_UNIFICADO):
            print("❌ Arquivo de referência ou unificado não encontrado.")
            return
        print(f"📂 Usando arquivo de referência: {ARQ_REFERENCIA}")
        df_ref = pd.read_excel(ARQ_REFERENCIA)
        df_uni = pd.read_excel(ARQ_UNIFICADO)
        # coluna temporária apenas para comparação (NÃO será salva)
        df_uni['Codigo6'] = df_uni['Código'].astype(str).str[:6]
        resultados = []
        for _, linha in df_ref.iterrows():
            descr = linha.get('DESCRIÇÃO', '')
            qtd = linha.get('QTD', 0)
            codigos_raw = str(linha.get('REFERÊNCIA', ''))
            codigos = [c.strip()[:6] for c in codigos_raw.split() if len(c.strip()) >= 6]
            for cod in codigos:
                match = df_uni[df_uni['Codigo6'] == cod]
                if not match.empty:
                    for _, item in match.iterrows():
                        valor_unitario = converter_valor(item['Valor Unitário'])
                        if valor_unitario is None:
                            continue
                        try:
                            qtd_float = float(str(qtd).replace(',', '.'))
                        except:
                            qtd_float = 0.0
                        valor_total = round(valor_unitario * qtd_float, 2)
                        resultados.append({
                            'Fornecedor': item['Fornecedor'],
                            'Código': item['Código'],
                            'Descrição': item['Descrição'],
                            'Valor Unitário': valor_unitario,
                            'QTD': qtd_float,
                            'Valor Total': valor_total
                        })
                    break 
        if resultados:
            df_resultado = pd.DataFrame(resultados)

            # coluna temporária para agrupar (será removida antes de salvar)
            df_resultado['Codigo6'] = df_resultado['Código'].astype(str).str[:6]
            df_resultado['Valor Unitário'] = pd.to_numeric(df_resultado['Valor Unitário'], errors='coerce')

            idx_min = df_resultado.groupby('Codigo6')['Valor Unitário'].idxmin()
            df_dedup = df_resultado.loc[idx_min].reset_index(drop=True)

            # Recalcular Valor Total (garantia)
            df_dedup['Valor Total'] = (df_dedup['Valor Unitário'] * pd.to_numeric(df_dedup['QTD'], errors='coerce')).round(2)
            if 'Codigo6' in df_dedup.columns:
                df_dedup = df_dedup.drop(columns=['Codigo6'])
            # Salva a tabela comparada deduplicada
            df_dedup.to_excel(ARQ_SAIDA, index=False)
            print(f"✅ Comparação concluída com deduplicação por Código(6) e sem a coluna 'Codigo6' no arquivo final.")
            print(f"📁 Arquivo gerado: {ARQ_SAIDA}")

            # ⇩ Gerar tabelas por fornecedor (também SEM Codigo6)
            dir_Fornecedores = 'relatorio/Fornecedores'
            os.makedirs(dir_Fornecedores, exist_ok=True)
            for fornecedor, df_forn in df_dedup.groupby('Fornecedor'):
                safe_fornecedor = re.sub(r'[^A-Za-z0-9_]+', '_', (fornecedor or 'DESCONHECIDO')).strip('_')
                out_path = os.path.join(dir_Fornecedores, f"itens_{safe_fornecedor}.xlsx")
                df_forn = df_forn.sort_values(['Código', 'Valor Unitário']).reset_index(drop=True)
                # >>> ADICIONE ESTA LINHA <<<
                df_forn = df_forn.drop(columns=['Codigo6'], errors='ignore')
                # >>> FIM DA ADIÇÃO <<<
                df_forn.to_excel(out_path, index=False)
                print(f"📄 Tabela do fornecedor '{fornecedor}' salva em: {out_path}")
        else:
            print("⚠️ Nenhum código encontrado.")

    # Executa as funções
    comparar_codigos_e_gerar_resultado()

# ======= Variáveis globais para controlar o estado dos botões =======
arquivo_selecionado = False
referencia_selecionada = False
arquivo_paths = []  # Lista de caminhos dos arquivos
referencia_paths = []  # Lista de caminhos das referências
arquivo_tipos = []  # Lista de tipos dos arquivos
referencia_tipos = []  # Lista de tipos das referências
def criar_pasta_referencia():
    pasta = "referencia"
    if not os.path.exists(pasta):
        os.makedirs(pasta)
        print(f"📁 Pasta criada: {pasta}")
    else:
        print(f"📂 Pasta já existe: {pasta}")
criar_pasta_referencia()
def selecionar_arquivo():
    """Função para selecionar múltiplos arquivos"""
    global arquivo_selecionado, arquivo_paths, arquivo_tipos

    filenames = filedialog.askopenfilenames(
        title="Selecionar Arquivos (PDF ou Excel)",
        filetypes=[
            ("Arquivos PDF", "*.pdf"),
            ("Arquivos Excel", "*.xlsx"),
            ("Todos os arquivos", "*.*")
        ]
    )
    if filenames:
        arquivo_selecionado = True
        arquivo_paths = list(filenames)
        arquivo_tipos = []

        pdf_count = 0
        excel_count = 0
        outros_count = 0
        os.makedirs("pdf", exist_ok=True)
        for filename in filenames:
            if filename.lower().endswith(".pdf"):
                arquivo_tipos.append("pdf")
                pdf_count += 1
                destino = os.path.join("pdf", os.path.basename(filename))
                shutil.copy2(filename, destino)
                print(f"📥 PDF copiado para pasta 'pdf': {destino}")
            elif filename.lower().endswith(".xlsx"):
                arquivo_tipos.append("xlsx")
                excel_count += 1
            else:
                arquivo_tipos.append("outro")
                outros_count += 1
        descricao = []
        if pdf_count > 0:
            descricao.append(f"📄 {pdf_count} PDF(s)")
        if excel_count > 0:
            descricao.append(f"📊 {excel_count} Excel(s)")
        if outros_count > 0:
            descricao.append(f"📁 {outros_count} outro(s)")
        arquivo_label.config(
            text=f"📁 {len(filenames)} arquivo(s): {', '.join(descricao)}"
        )
        verificar_botoes()
        print(f"✅ {len(filenames)} arquivo(s) selecionado(s): {filenames}")
        print(f"📊 Tipos: PDF={pdf_count}, Excel={excel_count}, Outros={outros_count}")

# Função para Seleção de Referências no APP
def selecionar_referencia():
    """Função para selecionar múltiplos arquivos de referência"""
    global referencia_selecionada, referencia_paths, referencia_tipos
    filenames = filedialog.askopenfilenames(
        title="Selecionar Arquivos de Referência (PDF ou Excel)",
        filetypes=[
            ("Arquivos PDF", "*.pdf"),
            ("Arquivos Excel", "*.xlsx"),
            ("Todos os arquivos", "*.*")
        ]
    )
    if filenames:
        os.makedirs('referencia', exist_ok=True)
        novos_caminhos = []
        for file in filenames:
            destino = os.path.join('referencia', os.path.basename(file))
            shutil.copy2(file, destino)
            novos_caminhos.append(destino)
        referencia_selecionada = True
        referencia_paths = list(novos_caminhos)
        referencia_tipos = []

        pdf_count = 0
        excel_count = 0
        outros_count = 0

        for filename in filenames:
            if filename.lower().endswith('.pdf'):
                referencia_tipos.append("pdf")
                pdf_count += 1
            elif filename.lower().endswith('.xlsx'):
                referencia_tipos.append("xlsx")
                excel_count += 1
            else:
                referencia_tipos.append("outro")
                outros_count += 1

        descricao = []
        if pdf_count > 0:
            descricao.append(f"📄 {pdf_count} PDF(s)")
        if excel_count > 0:
            descricao.append(f"📊 {excel_count} Excel(s)")
        if outros_count > 0:
            descricao.append(f"📁 {outros_count} outro(s)")

        referencia_label.config(text=f"📋 {len(filenames)} referência(s): {', '.join(descricao)}")
        verificar_botoes()
        print(f"✅ {len(filenames)} referência(s) selecionada(s): {filenames}")
        print(f"📊 Tipos: PDF={pdf_count}, Excel={excel_count}, Outros={outros_count}")

# Função para Executar Análise
def executar_analise():
    def apagar_pastas_menos_relatorio():
        pastas_preservadas = {"assets","relatorio"}
        diretorio_atual = os.getcwd()
        for nome in os.listdir(diretorio_atual):
            caminho = os.path.join(diretorio_atual, nome)
            if os.path.isdir(caminho) and nome not in pastas_preservadas:
                try:
                    shutil.rmtree(caminho)
                    print(f"🗑️ Pasta apagada: {nome}")
                except Exception as e:
                    print(f"❌ Erro ao apagar a pasta {nome}: {e}")
    os.makedirs("pdf", exist_ok=True)

    """Função para executar análise com temporizador"""
    global btn_executar_analise, status_label, root
    btn_executar_analise.config(state="disabled", text="⏳ Executando...")
    status_label.config(text="🔍 Iniciando análise...")
    root.update_idletasks()
    inicio = time.time()
    try:
        if checkbox_com_ref_var.get():
            status_label.config(text="⚙️ Processando PDFs...")
            root.update_idletasks()
            sem_referencia()

            status_label.config(text="⚙️ Processando referências...")
            root.update_idletasks()
            tratar_referencias()

            status_label.config(text="⚙️ Localizando Dados...")
            root.update_idletasks()
            busca_codigo()
            apagar_pastas_menos_relatorio()
        else:
            status_label.config(text="⚙️ Executando sem referência...")
            root.update_idletasks()
            sem_referencia()
            apagar_pastas_menos_relatorio()
        fim = time.time()
        duracao = fim - inicio
        status_label.config(text=f"✅ Análise finalizada em {duracao:.2f} segundos.")
        btn_executar_analise.config(state="normal", text="🚀 Executar Análise")
    except Exception as e:
        status_label.config(text=f"❌ Erro na análise: {str(e)}")
        btn_executar_analise.config(state="normal", text="Tentar novamente")

# ======== Funções para Marcar/Desmarcar Checkboxes e Botões ========
def on_checkbox_com_referencia():
    """Função chamada quando checkbox 'Com Referência' é marcado/desmarcado"""
    global arquivo_selecionado, referencia_selecionada, arquivo_paths, referencia_paths, arquivo_tipos, referencia_tipos
    if checkbox_com_ref_var.get():
        checkbox_sem_ref_var.set(False)
    arquivo_selecionado = False
    referencia_selecionada = False
    arquivo_paths = []
    referencia_paths = []
    arquivo_tipos = []
    referencia_tipos = []
    arquivo_label.config(text="📁 Nenhum arquivo selecionado")
    referencia_label.config(text="📋 Nenhuma referência selecionada")
    verificar_botoes()

def on_checkbox_sem_referencia():
    """Função chamada quando checkbox 'Sem Referência' é marcado/desmarcado"""
    global arquivo_selecionado, referencia_selecionada, arquivo_paths, referencia_paths, arquivo_tipos, referencia_tipos
    if checkbox_sem_ref_var.get():
        checkbox_com_ref_var.set(False)
    arquivo_selecionado = False
    referencia_selecionada = False
    arquivo_paths = []
    referencia_paths = []
    arquivo_tipos = []
    referencia_tipos = []
    arquivo_label.config(text="📁 Nenhum arquivo selecionado")
    referencia_label.config(text="📋 Nenhuma referência selecionada")
    verificar_botoes()

def verificar_botoes():
    """Função para verificar e ativar/desativar botões baseado no estado"""
    com_ref_marcado = checkbox_com_ref_var.get()
    sem_ref_marcado = checkbox_sem_ref_var.get()
    if com_ref_marcado or sem_ref_marcado:
        btn_selecionar_arquivo.config(state="normal")
    else:
        btn_selecionar_arquivo.config(state="disabled")
    if com_ref_marcado:
        btn_selecionar_referencia.config(state="normal")
    else:
        btn_selecionar_referencia.config(state="disabled")
    if arquivo_selecionado and (not com_ref_marcado or referencia_selecionada):
        btn_executar_analise.config(state="normal")
    else:
        btn_executar_analise.config(state="disabled")

# ======== Função principal do APP ========
def main():
    # Declarações globais
    global btn_selecionar_arquivo, btn_selecionar_referencia, btn_executar_analise
    global arquivo_label, referencia_label
    global checkbox_com_ref_var, checkbox_sem_ref_var
    global root, status_label
    # Criar a janela principal com tema escuro do Bootstrap
    root = ttk.Window(
        title="Relátorio de Orçamentos",
        themename="darkly",
        size=(800, 700),
        resizable=(True, True)
    )
    try:
        root.iconbitmap("assets/icone.ico")
    except:
        pass
    main_frame = ttk.Frame(root, padding=20)
    main_frame.pack(fill=BOTH, expand=True)
    title_label = ttk.Label(
        main_frame,
        text="Relátorio de Orçamentos",
        bootstyle="primary",
        font=("Helvetica",25,"bold")
    )
    title_label.pack(pady=(0, 20))
    subtitle_label = ttk.Label(
        main_frame,
        text=" Automatizar o tratamento de orçamentos, limpando e estruturando os dados para gerar arquivos organizados em PDF.",
        bootstyle="secondary",
        foreground="#DEDEDE"
    )
    subtitle_label.pack(pady=(0, 30))
    checkbox_frame = ttk.LabelFrame(main_frame, text="📋 Opções de Análise", padding=15)
    checkbox_frame.pack(fill=X, pady=20)
    checkbox_com_ref_var = tk.BooleanVar()
    checkbox_sem_ref_var = tk.BooleanVar()
    checkbox_com_ref = ttk.Checkbutton(
        checkbox_frame,
        text="Com Referência",
        variable=checkbox_com_ref_var,
        command=on_checkbox_com_referencia,
        bootstyle="primary-round-toggle"
    )
    checkbox_com_ref.pack(anchor=W, pady=5)
    checkbox_sem_ref = ttk.Checkbutton(
        checkbox_frame,
        text="Sem Referência",
        variable=checkbox_sem_ref_var,
        command=on_checkbox_sem_referencia,
        bootstyle="primary-round-toggle"
    )
    checkbox_sem_ref.pack(anchor=W, pady=5)
    arquivo_frame = ttk.LabelFrame(main_frame, text="📁 Seleção de Arquivos", padding=15)
    arquivo_frame.pack(fill=X, pady=20)
    btn_selecionar_arquivo = ttk.Button(
        arquivo_frame,
        text="📁 Selecionar Arquivos",
        command=selecionar_arquivo,
        bootstyle="info-outline",
        width=25,
        state="disabled"
    )
    btn_selecionar_arquivo.pack(pady=5)
    arquivo_label = ttk.Label(
        arquivo_frame,
        text="📁 Nenhum arquivo selecionado",
        bootstyle="secondary"
    )
    arquivo_label.pack(pady=5)
    btn_selecionar_referencia = ttk.Button(
        arquivo_frame,
        text="📋 Selecionar Referência",
        command=selecionar_referencia,
        bootstyle="warning-outline",
        width=25,
        state="disabled"
    )
    btn_selecionar_referencia.pack(pady=5)
    referencia_label = ttk.Label(
        arquivo_frame,
        text="📋 Nenhuma referência selecionada",
        bootstyle="secondary"
    )
    referencia_label.pack(pady=5)
    execucao_frame = ttk.Frame(main_frame)
    execucao_frame.pack(pady=20)
    btn_executar_analise = ttk.Button(
        execucao_frame,
        text="🚀 Executar Análise",
        command=executar_analise,
        bootstyle="success-outline",
        width=30,
        state="disabled"
    )
    btn_executar_analise.pack(pady=10)
    btn_limpar = ttk.Button(
        execucao_frame,
        text="🧹 Limpar Selecionados",
        command=limpar_selecionados,
        bootstyle="secondary-outline",
        width=30
    )
    btn_limpar.pack(pady=5)
    btn_exportar = ttk.Button(
        execucao_frame,
        text="📦 Exportar",
        command=exportar_relatorio,
        bootstyle="primary-outline",
        width=30
    )
    btn_exportar.pack(pady=5)
    exit_button = ttk.Button(
        execucao_frame,
        text="❌ Sair",
        command=root.quit,
        bootstyle="danger-outline",
        width=30
    )
    exit_button.pack(pady=5)
    status_frame = ttk.Frame(root)
    status_frame.pack(side=BOTTOM, fill=X)
    status_label = ttk.Label(
        status_frame,
        text="✅ Sistema pronto para uso",
        bootstyle="success"
    )
    status_label.pack(side=LEFT, padx=10, pady=5)
    try:
        root.center_window()
    except:
        root.update_idletasks()
        width = root.winfo_width()
        height = root.winfo_height()
        x = (root.winfo_screenwidth() // 2) - (width // 2)
        y = (root.winfo_screenheight() // 2) - (height // 2)
        root.geometry(f'{width}x{height}+{x}+{y}')
    root.mainloop()

if __name__ == "__main__":
    main()